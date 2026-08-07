import os
import sys
import json
import datetime
import openpyxl
import pandas as pd
import streamlit as st

# ==================== إعدادات الصفحة والتصميم ====================
st.set_page_config(
    page_title="برنامج بودى للمشورة الأسرية",
    page_icon="🌸",
    layout="wide",
    initial_sidebar_state="expanded"
)

# تطبيق تنسيقات CSS مخصصة لتلائم الواجهة الوردية والأنيقة
st.markdown("""
    <style>
    .main {
        background-color: #FFF5F8;
    }
    .stButton>button {
        background-color: #EC4899;
        color: white;
        border-radius: 8px;
        font-weight: bold;
        border: none;
        padding: 0.5rem 1rem;
    }
    .stButton>button:hover {
        background-color: #BE185D;
        color: white;
    }
    h1, h2, h3 {
        color: #701A75;
    }
    </style>
""", unsafe_allow_html=True)

# ==================== الثوابت والملفات ====================
EXCEL_FILE = "template.xlsx"
USERS_FILE = "users.json"

DEFAULT_USERS = {
    "admin": {"pass": "admin123", "role": "admin", "name": "د. شيماء 🌸"},
    "user1": {"pass": "1234", "role": "user", "name": "د. علا 🎀"},
    "user2": {"pass": "1234", "role": "user", "name": "د. عبير 🎀"},
    "user3": {"pass": "1234", "role": "user", "name": "د. ايه 🎀"}
}

VISIT_SCHEDULE_OPTIONS = [
    "الاسبوع الاول", "عمر شهرين", "عمر 4 شهور", "عمر 6 شهور",
    "عمر 9 شهور", "عمر 12 شهر", "عمر 18 شهر", "عمر سنتين",
    "عمر سنتين ونصف", "عمر 3 سنين", "عمر 3 سنين ونصف", "عمر 4 سنين",
    "عمر 4 سنين ونصف", "عمر 5 سنين", "عمر 5 سنين ونصف", "عمر 6 سنين"
]

VISIT_SCHEDULE_MONTHS = {
    "الاسبوع الاول": 0.25, "عمر شهرين": 2, "عمر 4 شهور": 4, "عمر 6 شهور": 6,
    "عمر 9 شهور": 9, "عمر 12 شهر": 12, "عمر 18 شهر": 18, "عمر سنتين": 24,
    "عمر سنتين ونصف": 30, "عمر 3 سنين": 36, "عمر 3 سنين ونصف": 42, "عمر 4 سنين": 48,
    "عمر 4 سنين ونصف": 54, "عمر 5 سنين": 60, "عمر 5 سنين ونصف": 66, "عمر 6 سنين": 72
}

AUTO_FILL_DONE_FIELDS = [
    "فوائد الرضاعة الطبيعية والأوضاع وعلامات الجوع والشبع",
    "كفاية اللبن وكمية البراز",
    "إعطاء الجرعة اليومية من فيتامين د",
    "كيفية رعاية السرة والإهتمام بنظافة الطفل",
    "البطاقة الصحية وأهمية المتابعة الدورية ومنحنيات النمو",
    "أهمية الإلتزام بتطعيمات الطفل",
    "التغذية الصحية للأم المرضعة",
    "كيفية التعرف على علامات الخطورة"
]

PREGNANT_AUTO_FILL_DONE_FIELDS = [
    "التغذية السليمة", "المكملات الغذائية", "التمرينات الرياضية", "قسط من النوم والراحة",
    "المتابعة الدورية للحمل", "التحذير من تناول الأدوية بدون إستشارة طبيب والتعرض للتدخين والأبخرة",
    "المتاعب البسيطة في الشهور الأولى", "المتاعب في الشهور الأخيرة", "علامات الخطر أثناء الحمل",
    "مشاكل الولادة المبكرة وكيفية تجنبها", "حركة الجنين / معرفة جنس الجنين/ تمييز الأصوات من قبل الجنين",
    "تغير لون الجلد حول الحلمة وظهور بعض إفرازات من الثدي", "إرتداء الملابس الفضفاضة المريحة",
    "الإستعداد للولادة / تحضير ملابس المولود ، الخ", "علامات الولادة", "مميزات الولادة الطبيعية",
    "الساعة الذهبية الأولى", "ملامسة الجلد للجلد", "البداية المبكرة للرضاعة الطبيعية", "الرضاعة الطبيعية المطلقة",
    "أهمية المباعدة", "وسائل تنظيم الأسرة", "إستخدام وسيلة بعد الولادة مباشرة", "التطور العصبي والنفسي للطفل",
    "ملاحظات/ توصيات"
]

DROPDOWN_OPTIONS = {
    "مستوى التعليم": ["امى", "يجيد القراءة", "مؤهل متوسط", "فوق متوسط", "مؤهل عالى"],
    "مستوى التعليم للام": ["امى", "يجيد القراءة", "مؤهل متوسط", "فوق متوسط", "مؤهل عالى"],
    "مستوى التعليم للاب": ["امى", "يجيد القراءة", "مؤهل متوسط", "فوق متوسط", "مؤهل عالى"],
    "الوظيفة": ["يعمل", "لا تعمل"],
    "الوظيفة للام": ["يعمل", "لا تعمل"],
    "قرابة بين الزوجين": ["لا يوجد", "من الدرجة الأولي", "من الدرجة الثانية", "من الدرجة الخامسة"],
    "نوع الولادة": ["طبيعى", "قيصرى"],
    "مكان الولادة": ["المستشفى", "المنزل"],
    "وسيلة تنظيم الأسرة المستخدمة سابقا": ["لا يوجد", "اقراص", "حقن", "كبسولات", "لولب", "طرق طبيعية"],
    "وس وسائل تنظيم الأسرة": ["لا يوجد", "اقراص", "حقن", "كبسولات", "لولب", "طرق طبيعية"],
    "سبب دخول الحضانة": [
        "انخفاض وزن الطفل.", "احتياج الطفل لأدوية محددة بهذا الوقت.", "صعوبة شديدة في التنفس لعدم اكتمال نمو الرئتين.",
        "ارتفاع درجة حرارة جسم الرضيع.", "تعطل العمليات الحيوية بجسم الطفل.", "انخفاض معدل الجلوكوز في دم الطفل.",
        "معاناة الرضيع مشكلات في الجهاز الهضمي.", "إصابة الطفل بعدوى في الدم.", "إصابة الطفل بالصفراء.",
        "حدوث مشكلات خلال الولادة “الولادة المتعسرة أو الحمل الحرج”.", "وجود عيب خلقي يمنع الطفل عن التنفس أو الرضاعة بشكل طبيعي."
    ],
    "مكان المتابعة": ["وحدة", "مستشفى", "أخرى"],
    "مصدر الاحالة": ["مستشفى الولادة", "عيادة خاصة", "عيادة التطعيمات", "نصيحة"],
    "موعد الزيارة": VISIT_SCHEDULE_OPTIONS,
    "رضاعة طبيعية مطلقة": ["3 شهور", "4 شهور", "6 شهور"],
    "رضاعة طبيعية مع سوائل وأعشاب": ["تم", "لم يتم"],
    "رضاعة طبيعية مع صناعي": ["تم", "لم يتم"],
    "رضاعة لبن صناعي": ["تم", "لم يتم"],
    "موقف إستخدام وسيلة تنظيم أسرة": ["توجد", "لا يوجد", "وسؤال الحمل الجديد", "مرغوب", "غير مرغوب", "وسؤال الخدمات غير الملباة", "حدث", "لم يحدث"],
    "الحمل الجديد": ["مرغوب", "غير مرغوب"],
    "دخول الحضانة": ["تم", "لم يتم"],
    "ملامسة الجلد فى الساعة الذهبية الأولى": ["تم", "لم يتم"],
    "الرضاعة الطبيعية فى الساعة الذهبية الأولى": ["تم", "لم يتم"],
    "أمراض مزمنة: إرتفاع ضغط الدم": ["يوجد", "لا يوجد"],
    "أمراض مزمنة: السكر": ["يوجد", "لا يوجد"],
    "أمراض مزمنة: إضطرابات الغدة": ["يوجد", "لا يوجد"],
    "أمراض مزمنة: الأنيميا": ["يوجد", "لا يوجد"],
    'مكملات "قبل": حمض الفوليك': ["يوجد", "لا يوجد"],
    'مكملات "قبل": الحديد': ["يوجد", "لا يوجد"],
    'مكملات "قبل": الكالسيوم': ["يوجد", "لا يوجد"],
    'مكملات "أثناء": حمض الفوليك': ["يوجد", "لا يوجد"],
    'مكملات "أثناء": الحديد': ["يوجد", "لا يوجد"],
    'مكملات "أثناء": الكالسيوم': ["يوجد", "لا يوجد"],
    "النمو والتطور الحركي": ["طبيعي", "متأخر", "يحتاج متابعة", "تم التوعية"],
    "التطور الإدراكي والمعرفي": ["طبيعي", "متأخر", "يحتاج متابعة", "تم التوعية"],
    "التطور اللغوي": ["طبيعي", "متأخر", "يحتاج متابعة", "تم التوعية"],
    "رسائل التربية الإيجابية": ["طبيعي", "متأخر", "يحتاج متابعة", "تم التوعية"],
    "الأنشطة التحفيزية": ["طبيعي", "متأخر", "يحتاج متابعة", "تم التوعية"],
    "التوعية عن التغذية التكميلية وسلامة الغذاء والتغذية السليمة": ["طبيعي", "متأخر", "يحتاج متابعة", "تم التوعية"],
    "إعطاء الجرعة اليومية من الحديد": ["يوجد", "لا يوجد"],
    "أهمية إستخدام وسيلة تنظيم أسرة وأهمية المباعدة": ["تم التوعية", "غير مهتمة", "مرفوض"],
    "الخدمات الغير ملباه": ["يوجد", "لا يوجد"]
}

PREGNANT_COLUMNS = [
    "تاريخ التسجيل", "اسم المستخدم", "الاسم", "العنوان", "الرقم القومى", "رقم الموبايل", "العمر الحالى", "السن عند الزواج", "السن عند الحمل الاول",
    "مستوى التعليم", "الوظيفة", "تاريخ اخر دورة شهرية", "قرابة بين الزوجين", "عدد مرات الحمل", "عدد مرات الاجهاض",
    "عدد الاطفال", "المدة بين اخر حملين", "نوع الولادة", "أمراض مزمنة: إرتفاع ضغط الدم", "أمراض مزمنة: السكر",
    "أمراض مزمنة: إضطرابات الغدة", "أمراض مزمنة: الأنيميا", "أمراض مزمنة: اخرى", 'مكملات "قبل": حمض الفوليك',
    'مكملات "قبل": الحديد', 'مكملات "قبل": الكالسيوم', 'مكملات "أثناء": حمض الفوليك', 'مكملات "أثناء": الحديد',
    'مكملات "أثناء": الكالسيوم', "وسيلة تنظيم الأسرة المستخدمة سابقا", "مدة إستخدام الوسيلة السابقة", "شهر الحمل",
    "التاريخ الزيارة", "التغذية السليمة", "المكملات الغذائية", "التمرينات الرياضية", "قسط من النوم والراحة",
    "المتابعة الدورية للحمل", "التحذير من تناول الأدوية بدون إستشارة طبيب والتعرض للتدخين والأبخرة",
    "المتاعب البسيطة في الشهور الأولى", "المتاعب في الشهور الأخيرة", "علامات الخطر أثناء الحمل",
    "مشاكل الولادة المبكرة وكيفية تجنبها", "حركة الجنين / معرفة جنس الجنين/ تمييز الأصوات من قبل الجنين",
    "تغير لون الجلد حول الحلمة وظهور بعض إفرازات من الثدي", "إرتداء الملابس الفضفاضة المريحة",
    "الإستعداد للولادة / تحضير ملابس المولود ، الخ", "علامات الولادة", "مميزات الولادة الطبيعية",
    "الساعة الذهبية الأولى", "ملامسة الجلد للجلد", "البداية المبكرة للرضاعة الطبيعية", "الرضاعة الطبيعية المطلقة",
    "أهمية المباعدة", "وسائل تنظيم الأسرة", "إستخدام وسيلة بعد الولادة مباشرة", "التطور العصبي والنفسي للطفل",
    "ملاحظات/ توصيات", "تخطيط الزيارة القادمة", "المتابعة ما بعد الولادة"
]

CHILD_COLUMNS = [
    "تاريخ التسجيل", "اسم المستخدم", "تاريخ اول زيارة", "رقم الحالة", "اسم الام", "الرقم القومى للام", "رقم الموبايل للام", "تاريخ ميلاد الام",
    "مستوى التعليم للام", "عدد الاطفال لدى الام", "المدة بين اخر حملين", "الوظيفة للام", "الرقم القومى للاب",
    "رقم الموبايل للاب", "اسم الاب", "مستوى التعليم للاب", "اسم الطفل", "تاريخ الميلاد للطفل", "العمر الحالى للطفل (شهور)",
    "العمر الرحمى للطفل (أسابيع)", "وحدة", "مستشفى", "أخرى", "مستشفى الولادة", "عيادة خاصة", "عيادة التطعيمات", "نصيحة", "نوع الولادة", "مكان الولادة", "وزن الطفل عند الولادة",
    "طول الطفل عند الولادة", "مقاس راس الطفل عند الولادة", "دخول الحضانة", "سبب دخول الحضانة", "مدة البقاء فى الحضانة",
    "ملامسة الجلد فى الساعة الذهبية الأولى", "الرضاعة الطبيعية فى الساعة الذهبية الأولى", "موعد الزيارة",
    "تاريخ الزيارة", "رضاعة طبيعية مطلقة", "رضاعة طبيعية مع سوائل وأعشاب", "رضاعة طبيعية مع صناعي",
    "رضاعة لبن صناعي", "الوزن (كجم)", "الطول (سم)", "محيط الرأس (سم)", "فوائد الرضاعة الطبيعية والأوضاع وعلامات الجوع والشبع",
    "كفاية اللبن وكمية البراز", "إعطاء الجرعة اليومية من فيتامين د", "كيفية رعاية السرة والإهتمام بنظافة الطفل",
    "البطاقة الصحية وأهمية المتابعة الدورية ومنحنيات النمو", "أهمية الإلتزام بتطعيمات الطفل", "التغذية الصحية للأم المرضعة",
    "كيفية التعرف على علامات الخطورة", "النمو والتطور الحركي", "التطور الإدراكي والمعرفي",
    "التطور اللغوي", "رسائل التربية الإيجابية", "الأنشطة التحفيزية", "التوعية عن التغذية التكميلية وسلامة الغذاء والتغذية السليمة",
    "إعطاء الجرعة اليومية من الحديد", "أهمية إستخدام وسيلة تنظيم أسرة وأهمية المباعدة", "موقف إستخدام وسيلة تنظيم أسرة",
    "الحمل الجديد", "الخدمات الغير ملباه", "ملاحظات/ توصيات", "تخطيط الزيارة القادمة"
]

# ==================== الوظائف المساعدة ====================
def load_users():
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_USERS, f, ensure_ascii=False, indent=4)
        return DEFAULT_USERS
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return DEFAULT_USERS

def save_users(users_data):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users_data, f, ensure_ascii=False, indent=4)

def format_excel_as_text(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.number_format = '@'
                    if cell.value is not None:
                        cell.value = str(cell.value)
        wb.save(file_path)
    except Exception:
        pass

def extract_dob_from_national_id(nid):
    nid = str(nid).strip()
    if len(nid) == 14 and nid.isdigit():
        century_digit = nid[0]
        year_digits = nid[1:3]
        month_digits = nid[3:5]
        day_digits = nid[5:7]
        full_year = ('19' if century_digit == '2' else '20') + year_digits
        try:
            dob = datetime.date(int(full_year), int(month_digits), int(day_digits))
            return dob.strftime("%Y-%m-%d")
        except ValueError:
            return ""
    return ""

def evaluate_growth(age_months, current_weight, current_height):
    try:
        age = float(str(age_months).replace("شهر", "").strip())
        w = float(current_weight)
        h = float(current_height)
    except (ValueError, TypeError):
        return "طبيعي"
    
    if w < 2.5 or h < 45.0:
        return "متأخر"
    elif w > 15.3 or h > 95.0:
        return "متقدم"
    return "طبيعي"

def fetch_mother_history(national_id):
    national_id = str(national_id).strip()
    if not national_id or not os.path.exists(EXCEL_FILE):
        return {}
    try:
        data = {}
        excel = pd.ExcelFile(EXCEL_FILE)
        if "سجل المشورة للاطفال" in excel.sheet_names:
            df_c = pd.read_excel(excel, "سجل المشورة للاطفال", dtype=str)
            if "الرقم القومى للام" in df_c.columns:
                match_c = df_c[df_c["الرقم القومى للام"].astype(str).str.strip() == national_id]
                if not match_c.empty:
                    row_c = match_c.iloc[-1]
                    for col in df_c.columns:
                        data[col] = row_c.get(col, "")
        return data
    except Exception:
        return {}

# ==================== إدارة الجلسة وتسجيل الدخول ====================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user = None
    st.session_state.role = None
    st.session_state.name = None

users = load_users()

if not st.session_state.logged_in:
    st.markdown("<h2 style='text-align: center; color: #BE185D;'>🌸 برنامج بودى للمشورة الأسرية 🌸</h2>", unsafe_allow_html=True)
    st.markdown("<h4 style='text-align: center; color: #701A75;'>تسجيل الدخول</h4>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        user_options = {f"{v['name']} ({k})": k for k, v in users.items()}
        selected_display = st.selectbox("اختر الطبيبة / الحساب", list(user_options.keys()))
        username = user_options[selected_display]
        password = st.text_вища("كلمة المرور", type="password") if hasattr(st, 'text_вища') else st.text_input("كلمة المرور", type="password")
        
        if st.button("تسجيل الدخول ✨", use_container_width=True):
            if users[username]["pass"] == password:
                st.session_state.logged_in = True
                st.session_state.user = username
                st.session_state.role = users[username]["role"]
                st.session_state.name = users[username]["name"]
                st.rerun()
            else:
                st.error("كلمة المرور غير صحيحة!")
    st.stop()

# ==================== الشاشة الرئيسية والتنقل ====================
st.sidebar.markdown(f"### أهلاً بكِ د. {st.session_state.name} 🌸")
menu = st.sidebar.radio("القائمة الرئيسية", ["الصفحة الرئيسية", "سجل الحوامل", "سجل الأطفال", "لوحة الإحصائيات (Dashboard)", "إدارة المستخدمين"])

if st.sidebar.button("🚪 تسجيل الخروج"):
    st.session_state.logged_in = False
    st.rerun()

# ==================== 1. الصفحة الرئيسية ====================
if menu == "الصفحة الرئيسية":
    st.markdown("<h1>✨ مرحباً بكِ في نظام المشورة الأسرية ✨</h1>", unsafe_allow_html=True)
    st.write("استخدمي القائمة الجانبية للتنقل بين سجلات الحوامل، الأطفال، أو لوحة الإحصائيات.")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("🤰 الانتقال إلى سجل الحوامل", use_container_width=True):
            st.rerun()
    with col2:
        if st.button("👶 الانتقال إلى سجل الأطفال", use_container_width=True):
            st.rerun()

# ==================== 2. سجل الحوامل ====================
elif menu == "سجل الحوامل":
    st.markdown("<h2>🤰 سجل المشورة الأسرية للحوامل</h2>", unsafe_allow_html=True)
    
    form_data = {}
    with st.form("pregnant_form"):
        cols = st.columns(3)
        for idx, col_name in enumerate(PREGNANT_COLUMNS):
            if col_name in ["تاريخ التسجيل", "اسم المستخدم"]:
                continue
            with cols[idx % 3]:
                if col_name in DROPDOWN_OPTIONS:
                    form_data[col_name] = st.selectbox(col_name, DROPDOWN_OPTIONS[col_name])
                elif col_name in PREGNANT_AUTO_FILL_DONE_FIELDS:
                    form_data[col_name] = st.selectbox(col_name, ["تم", "لم يتم"], index=0)
                else:
                    form_data[col_name] = st.text_input(col_name)
        
        submitted = st.form_submit_button("💾 حفظ وتصدير للإكسيل")
        if submitted:
            form_data["تاريخ التسجيل"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            form_data["اسم المستخدم"] = st.session_state.user
            
            new_df = pd.DataFrame([form_data], dtype=str)
            if not os.path.exists(EXCEL_FILE):
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                    new_df.to_excel(writer, sheet_name="المشورة الاسرية للحامل", index=False)
            else:
                excel = pd.ExcelFile(EXCEL_FILE)
                all_dfs = {s: pd.read_excel(excel, sheet_name=s, dtype=str) for s in excel.sheet_names}
                if "المشورة الاسرية للحامل" in all_dfs:
                    all_dfs["المشورة الاسرية للحامل"] = pd.concat([all_dfs["المشورة الاسرية للحامل"], new_df], ignore_index=True)
                else:
                    all_dfs["المشورة الاسرية للحامل"] = new_df
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                    for s, df in all_dfs.items():
                        df.to_excel(writer, sheet_name=s, index=False)
            format_excel_as_text(EXCEL_FILE)
            st.success("تم حفظ البيانات وتصديرها للإكسيل بنجاح! ✨")

# ==================== 3. سجل الأطفال ====================
elif menu == "سجل الأطفال":
    st.markdown("<h2>👶 سجل المشورة الأسرية للأطفال</h2>", unsafe_allow_html=True)
    
    form_data = {}
    with st.form("child_form"):
        cols = st.columns(3)
        for idx, col_name in enumerate(CHILD_COLUMNS):
            if col_name in ["تاريخ التسجيل", "اسم المستخدم", "وحدة", "مستشفى", "أخرى", "مستشفى الولادة", "عيادة خاصة", "عيادة التطعيمات", "نصيحة"]:
                continue
            with cols[idx % 3]:
                if col_name in DROPDOWN_OPTIONS:
                    form_data[col_name] = st.selectbox(col_name, DROPDOWN_OPTIONS[col_name])
                elif col_name in AUTO_FILL_DONE_FIELDS:
                    form_data[col_name] = st.selectbox(col_name, ["تم", "لم يتم"], index=0)
                else:
                    form_data[col_name] = st.text_input(col_name)
        
        submitted = st.form_submit_button("💾 حفظ وتصدير للإكسيل للأطفال")
        if submitted:
            form_data["تاريخ التسجيل"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            form_data["اسم المستخدم"] = st.session_state.user
            
            new_df = pd.DataFrame([form_data], dtype=str)
            if not os.path.exists(EXCEL_FILE):
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                    new_df.to_excel(writer, sheet_name="سجل المشورة للاطفال", index=False)
            else:
                excel = pd.ExcelFile(EXCEL_FILE)
                all_dfs = {s: pd.read_excel(excel, sheet_name=s, dtype=str) for s in excel.sheet_names}
                if "سجل المشورة للاطفال" in all_dfs:
                    all_dfs["سجل المشورة للاطفال"] = pd.concat([all_dfs["سجل المشورة للاطفال"], new_df], ignore_index=True)
                else:
                    all_dfs["سجل المشورة للاطفال"] = new_df
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                    for s, df in all_dfs.items():
                        df.to_excel(writer, sheet_name=s, index=False)
            format_excel_as_text(EXCEL_FILE)
            st.success("تم حفظ البيانات وتصديرها للإكسيل بنجاح! ✨")

# ==================== 4. لوحة الإحصائيات ====================
elif menu == "لوحة الإحصائيات (Dashboard)":
    st.markdown("<h2>📊 لوحة البيانات والإحصائيات</h2>", unsafe_allow_html=True)
    if os.path.exists(EXCEL_FILE):
        excel = pd.ExcelFile(EXCEL_FILE)
        for s in excel.sheet_names:
            st.subheader(f"شيت: {s}")
            df = pd.read_excel(excel, sheet_name=s, dtype=str)
            st.dataframe(df)
    else:
        st.info("لا توجد بيانات مسجلة حتى الآن.")

# ==================== 5. إدارة المستخدمين ====================
elif menu == "إدارة المستخدمين":
    st.markdown("<h2>⚙️ إدارة حسابات الطبيبات</h2>", unsafe_allow_html=True)
    if st.session_state.role == "admin":
        selected_user = st.selectbox("اختر الطبيبة لتغيير كلمة المرور", list(users.keys()))
        new_pass = st.text_input("كلمة المرور الجديدة", type="password")
        if st.button("حفظ كلمة المرور"):
            users[selected_user]["pass"] = new_pass
            save_users(users)
            st.success("تم تحديث كلمة المرور بنجاح!")
    else:
        st.error("عذراً، هذه الصفحة مخصصة لمدير النظام (Admin) فقط.")
